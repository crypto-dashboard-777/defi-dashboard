#!/usr/bin/env python3
"""
Wallet Positions Refresh — Rabby API edition with history-aware dashboard.

Pulls every complex DeFi position for the wallet from api.rabby.io
(unauthenticated, same data plane as DeBank, full per-vault granularity
including health_rate). Filters to positions >$50, appends a row-per-position
to a persistent .xlsx log, saves a timestamped snapshot JSON, and re-renders
a self-contained HTML dashboard that shows current state plus deltas at
12h / 24h / 48h / 72h / 168h lookbacks for each position.

Designed to be invoked by a scheduled task. Usage:

  refresh_positions.py [--xlsx /path/to/log.xlsx]
                       [--snapshots-dir /path/to/snapshots/]
                       [--dashboard /path/to/dashboard/index.html]
                       [--out /path/to/latest.json]

Endpoints used:
  GET  https://api.rabby.io/v1/user/total_balance?id={addr}
  GET  https://api.rabby.io/v1/user/complex_protocol_list?id={addr}
"""
from __future__ import annotations
import argparse, json, os, sys, time, urllib.request, urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path

WALLET     = os.environ.get("DEBANK_WALLET", "0xe16be042f9433779909a972669be3a2003956348")
SOL_WALLET = os.environ.get("SOL_WALLET", "6QcRFrTcHCZgKdtX83iusXVBvcz3vrwiKayREZtJBx5o")
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Rabby/1.0"
FILTER_USD = 50.0
RABBY_BASE = "https://api.rabby.io/v1"
LOOKBACKS_H = [12, 24, 36, 48, 168, 336]   # 168 = 7d, 336 = 14d

BENCHMARK_DAYS = [
    (7,   "1w"),
    (14,  "2w"),
    (30,  "1mo"),
    (90,  "3mo"),
    (365, "1y"),
]
SOL_RPC = "https://api.mainnet-beta.solana.com"
SOL_MINT = "So11111111111111111111111111111111111111112"
TOKEN_PROGRAM = "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"
TOKEN_2022_PROGRAM = "TokenzQdBNbLqP5VEhdkAS6EPFLC1PHnBqCXEpPxuEb"

# Rabby chain IDs → human-readable display names
CHAIN_NAMES = {
    "eth": "Ethereum", "arb": "Arbitrum", "op": "Optimism", "matic": "Polygon",
    "base": "Base", "bsc": "BNB Chain", "avax": "Avalanche", "ftm": "Fantom",
    "mnt": "Mantle", "scrl": "Scroll", "linea": "Linea", "blast": "Blast",
    "zksync": "zkSync Era", "plasma": "Plasma", "celo": "Celo", "metis": "Metis",
    "era": "zkSync Era", "mantle": "Mantle",
}


def chain_display(chain_id: str) -> str:
    return CHAIN_NAMES.get(chain_id, chain_id.title() if chain_id else "?")


def rabby_get(path: str, params: dict | None = None, retries: int = 3, backoff: float = 1.5):
    qs = ""
    if params:
        from urllib.parse import urlencode
        qs = "?" + urlencode(params)
    url = f"{RABBY_BASE}{path}{qs}"
    req = urllib.request.Request(url, headers={"accept": "application/json", "User-Agent": UA})
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=25) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            if e.code in (429, 502, 503, 504) and attempt + 1 < retries:
                time.sleep(backoff ** (attempt + 1))
                continue
            raise
        except (urllib.error.URLError, TimeoutError):
            if attempt + 1 < retries:
                time.sleep(backoff ** (attempt + 1))
                continue
            raise


def rabby_total_balance() -> float:
    data = rabby_get("/user/total_balance", {"id": WALLET})
    return float(data.get("total_usd_value", 0))


def rabby_protocols():
    return rabby_get("/user/complex_protocol_list", {"id": WALLET})


def rabby_tokens():
    return rabby_get("/user/token_list", {"id": WALLET, "is_all": "true"})


def collect_position_token_ids(protocols) -> set:
    """Build set of (chain, token_id_lower) for tokens that are likely position WRAPPERS
    (LP tokens, PT tokens, aTokens) — i.e., tokens whose token_list balance is the
    same as the position itself (double-counted). Borrow tokens (USDT/USDC etc.)
    are NOT included since the user can also hold them plainly in the wallet."""
    out = set()
    for proto in protocols:
        for pf in proto.get("portfolio_item_list", []):
            det = pf.get("detail") or {}
            # supply side: usually wrappers (aTokens, LP shares, PT tokens) — exclude
            for key in ("supply_token_list", "asset_token_list", "token_list"):
                for t in (det.get(key) or []):
                    tid = (t.get("id") or "").lower()
                    chain = t.get("chain") or ""
                    if tid and chain:
                        out.add((chain, tid))
            for t in (pf.get("asset_token_list") or []):
                tid = (t.get("id") or "").lower()
                chain = t.get("chain") or ""
                if tid and chain:
                    out.add((chain, tid))
    return out


def wallet_tokens(tokens, position_token_ids: set, min_usd: float = FILTER_USD) -> list:
    """Plain wallet tokens not deployed in any position, value > min_usd.
    Filters out deposit-receipt / wrapper tokens by protocol_id."""
    # protocol_ids that indicate a stablecoin / native asset issuer (KEEP these)
    STABLE_ISSUER_PIDS = {
        "tether", "usd-coin", "circle", "paypal", "paypal-usd",
        "paxos-standard", "binance-usd", "frax", "true-usd", "dai",
        "ethereum", "weth", "wbtc", "bitcoin",
    }
    # Airdrop scam / illiquid junk tokens — blocked by contract address
    BLOCKED_IDS = {
        "0x66a3c2fa3e467aa586e90912f977e648589cabaf",  # AICC (AI Chain Coin) — airdrop scam, ~$0 liquidity
    }

    out = []
    for t in tokens or []:
        amount = float(t.get("amount") or 0)
        price = float(t.get("price") or 0)
        usd = amount * price
        if usd < min_usd:
            continue
        chain = t.get("chain") or "?"
        tid = (t.get("id") or "").lower()
        proto_id = (t.get("protocol_id") or "").lower()
        if tid in BLOCKED_IDS:
            continue
        # Drop position wrappers: token has a DeFi protocol_id (e.g., aave, pendle, morpho).
        # Plain assets (no protocol_id) and stablecoin issuers (tether, paypal) are kept.
        if proto_id and proto_id not in STABLE_ISSUER_PIDS:
            continue
        out.append({
            "chain": chain_display(chain),
            "chainId": chain,
            "token": t.get("symbol") or t.get("optimized_symbol") or t.get("display_symbol") or "?",
            "id": tid,
            "amount": amount,
            "usd": usd,
        })
    out.sort(key=lambda x: -x["usd"])
    return out


def wallet_token_key(t: dict) -> str:
    """Stable key per wallet token: chain + token id (or symbol fallback)."""
    return f"{t.get('chainId','?')}|{(t.get('id') or t.get('token','?')).lower()}"


def usd_of(tok: dict) -> float:
    return (tok.get("amount", 0) or 0) * (tok.get("price", 0) or 0)


def normalize(rabby_data) -> list[dict]:
    """Each portfolio_item in Rabby's response becomes one row."""
    out = []
    for proto in rabby_data:
        pname = proto.get("name") or "Unknown"
        chain_id = proto.get("chain") or "?"
        for pf in proto.get("portfolio_item_list", []):
            det = pf.get("detail") or {}
            sup_toks = det.get("supply_token_list") or []
            bor_toks = det.get("borrow_token_list") or []
            rew_toks = det.get("reward_token_list") or []
            if not sup_toks and not bor_toks and not rew_toks:
                generic = pf.get("asset_token_list") or det.get("token_list") or []
                if generic:
                    sup_toks = generic

            supplied = [{
                "token": t.get("symbol") or t.get("optimized_symbol") or "?",
                "amount": float(t.get("amount") or 0),
                "usd": float(usd_of(t)),
            } for t in sup_toks if usd_of(t) >= 0.01]
            borrowed = [{
                "token": t.get("symbol") or t.get("optimized_symbol") or "?",
                "amount": float(t.get("amount") or 0),
                "usd": float(usd_of(t)),
            } for t in bor_toks if usd_of(t) >= 0.01]
            rewards = [{
                "token": t.get("symbol") or t.get("optimized_symbol") or "?",
                "amount": float(t.get("amount") or 0),
                "usd": float(usd_of(t)),
            } for t in rew_toks if usd_of(t) >= 0.01]

            sup_sum = sum(l["usd"] for l in supplied)
            bor_sum = sum(l["usd"] for l in borrowed)
            rew_sum = sum(l["usd"] for l in rewards)
            net = sup_sum - bor_sum + rew_sum
            ptype = pf.get("name") or "Position"

            hr_raw = det.get("health_rate")
            hr = None
            if hr_raw is not None:
                try:
                    hrf = float(hr_raw)
                    if hrf > 100:
                        hr = None
                    else:
                        hr = round(hrf, 4)
                except (TypeError, ValueError):
                    hr = None

            pool = pf.get("pool") or {}
            pool_id = pool.get("id") or pool.get("controller") or ""
            description = det.get("description") or ""

            out.append({
                "protocol": pname,
                "chain": chain_display(chain_id),
                "chainId": chain_id,
                "type": "Yield" if not borrowed else "Lending",
                "rabbyType": ptype,
                "healthRate": hr,
                "supplied": supplied,
                "borrowed": borrowed,
                "rewards": rewards,
                "description": description,
                "_supSum": sup_sum,
                "_borSum": bor_sum,
                "_rewSum": rew_sum,
                "_net": net,
                "_poolId": pool_id,
                "_source": "rabby",
            })
    return out


def keep(p: dict) -> bool:
    return p["_supSum"] > FILTER_USD or p["_borSum"] > FILTER_USD or abs(p["_net"]) > FILTER_USD


def position_key(p: dict) -> str:
    """Stable, unique-per-market key including supply+borrow token signature.
    Works on both raw (with _poolId/_supSum) and rendered (with poolId/supSum) dicts.
    """
    pool = (p.get("_poolId") or p.get("poolId") or "")
    pool = pool[:14] if pool else "nopool"
    sup_syms = "+".join(sorted({l["token"] for l in p.get("supplied", [])})) or "none"
    bor_syms = "+".join(sorted({l["token"] for l in p.get("borrowed", [])})) or "none"
    return f"{p['protocol'].replace(' ', '')}|{p.get('chainId','?')}|{pool}|{sup_syms}|{bor_syms}"


def position_id(p: dict) -> str:
    """Short id for the xlsx log (kept for backward compat)."""
    if p.get("_poolId"):
        return f"{p['protocol'].replace(' ', '')}-{p['chainId']}-{p['_poolId'][:10]}"
    sup = "+".join(sorted(l["token"] for l in p["supplied"])) or "none"
    bor = "+".join(sorted(l["token"] for l in p["borrowed"])) or "none"
    return f"{p['protocol'].replace(' ', '')}-{p['chainId']}-{sup}~{bor}"


def position_name(p: dict) -> str:
    sup = "+".join(l["token"] for l in p["supplied"]) or "—"
    bor = "+".join(l["token"] for l in p["borrowed"]) or "(no debt)"
    return f"{sup} vs {bor}"


def fmt_tokens(rows):
    return "; ".join(f"{l['token']} {l['amount']:,.4f}" for l in rows) if rows else ""


# ────────────────────────── XLSX log ──────────────────────────

def append_to_xlsx(xlsx_path: Path, snapshot_time: str, source_label: str,
                   positions, totals, reported_total):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule

    headers_log = [
        "Snapshot (UTC)", "Wallet", "Chain", "Protocol", "Type",
        "Position ID", "Position", "Health Rate",
        "Supplied Tokens", "Supplied USD",
        "Borrowed Tokens", "Borrowed USD",
        "Net USD", "LTV", "Source",
    ]
    headers_sum = [
        "Snapshot (UTC)", "Wallet", "Reported Total USD",
        "Computed Net Equity USD", "Total Supplied USD", "Total Borrowed USD",
        "Position Count", "Lowest Health Rate", "Source",
    ]

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb["Positions Log"]
        ws2 = wb["Snapshot Summary"]
        new_workbook = False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Positions Log"
        ws.append(headers_log)
        ws2 = wb.create_sheet("Snapshot Summary")
        ws2.append(headers_sum)
        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="1F2937")
        for ws_, hdrs in [(ws, headers_log), (ws2, headers_sum)]:
            for i in range(len(hdrs)):
                c = ws_.cell(row=1, column=i+1)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_.freeze_panes = "A2"
        for col, w in dict(A=19, B=44, C=11, D=18, E=9, F=36, G=32, H=11,
                           I=38, J=14, K=38, L=14, M=14, N=8, O=26).items():
            ws.column_dimensions[col].width = w
        for col, w in dict(A=19, B=44, C=18, D=22, E=18, F=18, G=14, H=16, I=26).items():
            ws2.column_dimensions[col].width = w
        new_workbook = True

    usd_fmt = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    ltv_fmt = '0.0%;[Red]-0.0%;"-"'
    hr_fmt = '0.00'
    arial = Font(name="Arial", size=10)

    for p in positions:
        row_idx = ws.max_row + 1
        ws.append([
            snapshot_time, WALLET, p["chain"], p["protocol"], p["type"],
            position_id(p), position_name(p),
            p.get("healthRate") if p.get("healthRate") is not None else "",
            fmt_tokens(p["supplied"]), p["_supSum"],
            fmt_tokens(p["borrowed"]), p["_borSum"],
            f"=J{row_idx}-L{row_idx}",
            f"=IF(J{row_idx}=0,\"\",L{row_idx}/J{row_idx})",
            source_label,
        ])
        for col, fmt in [(8, hr_fmt), (10, usd_fmt), (12, usd_fmt), (13, usd_fmt), (14, ltv_fmt)]:
            ws.cell(row=row_idx, column=col).number_format = fmt
        for c in ws[row_idx]:
            if not c.font.bold:
                c.font = arial

    sum_row = ws2.max_row + 1
    ws2.append([
        snapshot_time, WALLET, reported_total or "",
        totals["net"], totals["sup"], totals["bor"],
        totals["count"], totals["lowest_hr"] if totals["lowest_hr"] is not None else "",
        source_label,
    ])
    for col, fmt in [(3, usd_fmt), (4, usd_fmt), (5, usd_fmt), (6, usd_fmt), (8, hr_fmt)]:
        ws2.cell(row=sum_row, column=col).number_format = fmt
    for c in ws2[sum_row]:
        if not c.font.bold:
            c.font = arial

    if new_workbook:
        red = PatternFill("solid", start_color="FDECEA")
        amber = PatternFill("solid", start_color="FFF4E0")
        green = PatternFill("solid", start_color="E8F5EC")
        rng = "H2:H10000"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["1.05"], fill=red))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["1.05", "1.15"], fill=amber))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["1.15"], fill=green))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return new_workbook


# ────────────────────────── Snapshot store ──────────────────────────

def snapshot_dict(positions, totals, reported_total, snapshot_time, source_label, wallet_toks=None):
    """Convert raw positions to the JSON snapshot shape (drop _-prefixed fields)."""
    wallet_toks = wallet_toks or []
    wallet_total = sum(t["usd"] for t in wallet_toks)
    return {
        "wallet": WALLET,
        "capturedAt": snapshot_time,
        "source": source_label,
        "reportedTotal": reported_total,
        "totals": {**totals, "walletTokensTotal": wallet_total},
        "positions": [
            {**{k: v for k, v in p.items() if not k.startswith("_") and k != "rewards"},
             "supSum": p["_supSum"], "borSum": p["_borSum"], "rewSum": p["_rewSum"],
             "net": p["_net"], "poolId": p["_poolId"]}
            for p in positions
        ],
        "walletTokens": wallet_toks,
    }


def save_timestamped_snapshot(snapshots_dir: Path, snap: dict) -> Path:
    """Write snap to snapshots/<timestamp>.json. Returns the path."""
    snapshots_dir.mkdir(parents=True, exist_ok=True)
    ts_safe = snap["capturedAt"].replace(":", "-").replace(" ", "T") + "Z"
    path = snapshots_dir / f"{ts_safe}.json"
    with open(path, "w") as f:
        json.dump(snap, f, separators=(",", ":"), default=str)
    # also keep latest.json (pretty) and latest.compact.json (compact) for ad-hoc inspection
    with open(snapshots_dir / "latest.json", "w") as f:
        json.dump(snap, f, indent=2, default=str)
    with open(snapshots_dir / "latest.compact.json", "w") as f:
        json.dump(snap, f, separators=(",", ":"), default=str)
    return path


def load_all_snapshots(snapshots_dir: Path) -> list[dict]:
    """Load all timestamped snapshot files, sorted by capturedAt ascending."""
    if not snapshots_dir.exists():
        return []
    snaps = []
    for f in snapshots_dir.iterdir():
        if not f.is_file() or not f.suffix == ".json":
            continue
        if f.name.startswith("latest"):
            continue
        try:
            with open(f) as fh:
                snaps.append(json.load(fh))
        except (json.JSONDecodeError, OSError):
            pass
    snaps.sort(key=lambda s: s.get("capturedAt", ""))
    return snaps


# ────────────────────────── Dashboard renderer ──────────────────────────

def parse_captured(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)


def find_lookback_snapshot(snapshots: list[dict], current_t: datetime, hours: int):
    """Given list of snapshots (sorted asc) and the current time, find the snapshot
    closest to (current_t - hours), within ±hours/2 tolerance.
    For the 12h slot specifically, fall back to the most-recent prior snapshot
    if none falls within tolerance (so the column populates as soon as ≥2 snapshots exist).
    Returns (snapshot_dict, hours_ago_actual) or (None, None)."""
    target = current_t - timedelta(hours=hours)
    tolerance = timedelta(hours=hours / 2)
    best = None
    best_dist = None
    for s in snapshots:
        try:
            t = parse_captured(s["capturedAt"])
        except (ValueError, KeyError):
            continue
        dist = abs(t - target)
        if dist <= tolerance and (best_dist is None or dist < best_dist):
            best = s
            best_dist = dist
    if best is None and hours == 12:
        # Fallback: use most recent prior snapshot at least 1h old (skip back-to-back
        # script runs so the column shows a meaningful baseline).
        min_age = timedelta(hours=1)
        prior = [s for s in snapshots
                 if s.get("capturedAt") and (current_t - parse_captured(s["capturedAt"])) >= min_age]
        if prior:
            best = prior[-1]
        else:
            # Last resort: any prior snapshot
            any_prior = [s for s in snapshots
                         if s.get("capturedAt") and parse_captured(s["capturedAt"]) < current_t]
            if any_prior:
                best = any_prior[-1]
    if best is None:
        return None, None
    actual_hours = (current_t - parse_captured(best["capturedAt"])).total_seconds() / 3600
    return best, actual_hours


def build_history(current: dict, snapshots: list[dict]):
    """For the current snapshot, attach a `history` dict to each position
    containing per-lookback comparison data, and produce a list of closed positions.
    Returns a *new* enriched snapshot dict."""
    current_t = parse_captured(current["capturedAt"])
    # exclude the current snapshot itself from the history pool
    older = [s for s in snapshots if s.get("capturedAt") != current["capturedAt"]]

    # Pre-compute which snapshot maps to each lookback — deduplicate so the
    # same snapshot never appears in two columns (avoids "24h = 36h" clones).
    _used_snaps: set[str] = set()
    _lookback_map: dict[int, tuple] = {}
    for h in LOOKBACKS_H:
        snap, hago = find_lookback_snapshot(older, current_t, h)
        if snap is None or snap.get("capturedAt") in _used_snaps:
            _lookback_map[h] = (None, None)
        else:
            _used_snaps.add(snap["capturedAt"])
            _lookback_map[h] = (snap, hago)

    cur_keys = {position_key(p): p for p in current["positions"]}

    # Per-position history
    enriched_positions = []
    for p in current["positions"]:
        key = position_key(p)
        history = {}
        for h in LOOKBACKS_H:
            snap, hago = _lookback_map[h]
            if snap is None:
                history[str(h)] = None
                continue
            match = next((q for q in snap["positions"] if position_key(q) == key), None)
            if match is None:
                history[str(h)] = {
                    "at": snap["capturedAt"],
                    "hoursAgo": round(hago, 1),
                    "absent": True,
                }
            else:
                history[str(h)] = {
                    "at": snap["capturedAt"],
                    "hoursAgo": round(hago, 1),
                    "supSum": match["supSum"],
                    "borSum": match["borSum"],
                    "net": match["net"],
                    "healthRate": match.get("healthRate"),
                    "deltaNet": p["net"] - match["net"],
                    "deltaSup": p["supSum"] - match["supSum"],
                    "deltaBor": p["borSum"] - match["borSum"],
                    "deltaHr": (p.get("healthRate") - match["healthRate"])
                                if (p.get("healthRate") is not None and match.get("healthRate") is not None) else None,
                }
        enriched = dict(p)
        enriched["history"] = history
        enriched_positions.append(enriched)

    # Closed/absent positions: any key in any older snapshot but not in current
    closed = {}
    for s in older:
        for q in s["positions"]:
            k = position_key(q)
            if k in cur_keys:
                continue
            entry = closed.get(k)
            if entry is None or s["capturedAt"] > entry["lastSeenAt"]:
                closed[k] = {
                    "key": k,
                    "lastSeenAt": s["capturedAt"],
                    "protocol": q["protocol"],
                    "chain": q["chain"],
                    "name": position_name(q),
                    "supSum": q["supSum"],
                    "borSum": q["borSum"],
                    "net": q["net"],
                    "healthRate": q.get("healthRate"),
                }

    # Wallet tokens history
    cur_wallet = current.get("walletTokens") or []
    enriched_wallet = []
    for wt in cur_wallet:
        key = wallet_token_key(wt)
        history = {}
        for h in LOOKBACKS_H:
            snap, hago = _lookback_map[h]
            if snap is None:
                history[str(h)] = None
                continue
            match = next((q for q in (snap.get("walletTokens") or []) if wallet_token_key(q) == key), None)
            if match is None:
                history[str(h)] = {"at": snap["capturedAt"], "hoursAgo": round(hago, 1), "absent": True}
            else:
                history[str(h)] = {
                    "at": snap["capturedAt"],
                    "hoursAgo": round(hago, 1),
                    "usd": match["usd"],
                    "amount": match["amount"],
                    "deltaUsd": wt["usd"] - match["usd"],
                    "deltaAmount": wt["amount"] - match["amount"],
                }
        e = dict(wt)
        e["history"] = history
        enriched_wallet.append(e)

    benchmarks = compute_benchmarks(current, snapshots)

    # Day 0 baseline — permanent file written once
    # Use net_pos (t.net + walletTokensTotal) for a consistent apples-to-apples comparison
    day0_info = None
    day0_path = Path(__file__).parent.parent / "outputs/snapshots/day0.json"
    if day0_path.exists():
        try:
            d0 = json.loads(day0_path.read_text())
            d0_t   = d0.get("totals") or {}
            d0_net_pos = (d0_t.get("net") or 0) + (d0_t.get("walletTokensTotal") or 0)
            # Fall back to reportedTotal if totals not present
            if d0_net_pos == 0:
                d0_net_pos = d0.get("reportedTotal") or 0
            cur_t  = current.get("totals") or {}
            cur_net_pos = (cur_t.get("net") or 0) + (cur_t.get("walletTokensTotal") or 0)
            if cur_net_pos == 0:
                cur_net_pos = current.get("reportedTotal") or 0
            delta = cur_net_pos - d0_net_pos
            pct   = round(delta / d0_net_pos * 100, 2) if d0_net_pos else 0
            day0_info = {
                "net_pos": round(d0_net_pos, 2),
                "equity":  round(d0_net_pos, 2),   # kept for compat
                "at":      d0.get("capturedAt"),
                "delta":   round(delta, 2),
                "pct":     pct,
            }
        except Exception:
            pass

    return {
        **current,
        "positions": enriched_positions,
        "walletTokens": enriched_wallet,
        "closedPositions": list(closed.values()),
        "snapshotCount": len(snapshots),
        "earliestSnapshotAt": snapshots[0]["capturedAt"] if snapshots else current["capturedAt"],
        "lookbacksHours": LOOKBACKS_H,
        "benchmarks": benchmarks,
        "day0": day0_info,
    }


def compute_benchmarks(current: dict, snapshots: list[dict]) -> list[dict]:
    """Compare current total equity against medium/long-term benchmarks.
    Uses a ±50 % time tolerance so the first match within half the window counts.
    Returns list of {label, days, equity, delta, pct, at}."""
    current_t  = parse_captured(current["capturedAt"])
    cur_equity = current.get("reportedTotal") or 0
    older = [s for s in snapshots
             if s.get("capturedAt") and s["capturedAt"] != current["capturedAt"]
             and parse_captured(s["capturedAt"]) < current_t]

    result = []
    for days, label in BENCHMARK_DAYS:
        target     = current_t - timedelta(days=days)
        tolerance  = timedelta(days=max(3, days * 0.5))
        candidates = [s for s in older
                      if abs(parse_captured(s["capturedAt"]) - target) <= tolerance]
        if candidates:
            best       = min(candidates, key=lambda s: abs(parse_captured(s["capturedAt"]) - target))
            old_equity = best.get("reportedTotal") or 0
            delta      = cur_equity - old_equity
            pct        = (delta / old_equity * 100) if old_equity else 0
            result.append({"label": label, "days": days,
                           "equity": round(old_equity, 2),
                           "delta":  round(delta, 2),
                           "pct":    round(pct, 2),
                           "at":     best["capturedAt"]})
        else:
            result.append({"label": label, "days": days,
                           "equity": None, "delta": None, "pct": None, "at": None})
    return result


DASHBOARD_CSS = """
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
:root {
  --bg:#05060c;
  --surface:#0b0e18;
  --surface-2:#111520;
  --surface-3:#171d2e;
  --border:#1a2035;
  --border-2:#242e48;
  --text:#e2e8f5;
  --muted:#6b7a96;
  --muted-2:#3d4d68;
  --green:#2dd4a0;
  --green-dim:rgba(45,212,160,.09);
  --green-glow:rgba(45,212,160,.2);
  --red:#f47070;
  --red-dim:rgba(244,112,112,.09);
  --amber:#f5a623;
  --amber-dim:rgba(245,166,35,.1);
  --indigo:#818cf8;
  --indigo-dim:rgba(129,140,248,.1);
  --supply:#2dd4a0;
  --borrow:#fb923c;
  --radius:12px;
}
*{box-sizing:border-box;margin:0;padding:0}
html,body{background:var(--bg);color:var(--text);
  font-family:'Inter',-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  font-size:13px;line-height:1.5;-webkit-font-smoothing:antialiased}
body{padding:24px 28px 80px;max-width:1440px;margin:0 auto}

/* ── Header ── */
header.page{display:flex;justify-content:space-between;align-items:flex-start;
  gap:20px;margin-bottom:20px;flex-wrap:wrap;
  background:linear-gradient(135deg,rgba(129,140,248,.04) 0%,transparent 60%);
  border:1px solid var(--border);border-radius:var(--radius);
  padding:16px 20px}
.hdr-left{display:flex;flex-direction:column;gap:4px}
.hdr-title{display:flex;align-items:center;gap:10px}
h1{font-size:18px;font-weight:800;letter-spacing:-.04em;color:var(--text)}
.live-dot{width:7px;height:7px;border-radius:50%;background:var(--green);flex-shrink:0;
  box-shadow:0 0 8px var(--green-glow);animation:pulse 2.5s ease-in-out infinite}
@keyframes pulse{0%{box-shadow:0 0 0 0 var(--green-glow)}
  70%{box-shadow:0 0 0 6px rgba(45,212,160,0)}100%{box-shadow:0 0 0 0 rgba(45,212,160,0)}}
.wallet{font-family:ui-monospace,"SFMono-Regular",Menlo,monospace;
  font-size:10px;color:var(--muted-2);letter-spacing:.02em}
.hdr-right{text-align:right;display:flex;flex-direction:column;gap:3px}
.hdr-right .snap-time{font-size:11px;color:var(--muted);font-weight:500}
.hdr-right .source{font-size:9.5px;color:var(--muted-2)}
.hdr-right .hist{font-size:9.5px;color:var(--muted-2);font-style:italic}
.live-line{font-size:11px;color:var(--muted);margin-top:3px}

/* ── Pills ── */
.pill{font-size:10px;padding:2px 8px;border-radius:999px;font-weight:700;
  white-space:nowrap;display:inline-block;letter-spacing:.02em}
.pill.green{color:#0ecf8a;background:rgba(14,207,138,.12);border:1px solid rgba(14,207,138,.22)}
.pill.amber{color:var(--amber);background:rgba(245,166,35,.12);border:1px solid rgba(245,166,35,.22)}
.pill.red{color:var(--red);background:rgba(244,112,112,.12);border:1px solid rgba(244,112,112,.22)}
.pill.gray{color:var(--muted);background:var(--surface-2);border:1px solid var(--border-2)}

/* ── Section labels ── */
.section-row{display:flex;justify-content:space-between;align-items:center;
  margin:20px 0 10px;flex-wrap:wrap;gap:8px}
.section-title{font-size:10px;color:var(--muted-2);text-transform:uppercase;
  letter-spacing:.1em;font-weight:700}
.legend{font-size:9.5px;color:var(--muted-2);display:flex;gap:14px;
  align-items:center;flex-wrap:wrap}
.legend .swatch{display:inline-block;width:6px;height:6px;border-radius:2px;
  margin-right:4px;vertical-align:-1px}

/* ── Main table ── */
.tbl{width:100%;border-collapse:separate;border-spacing:0;
  background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);overflow:hidden;
  box-shadow:0 1px 3px rgba(0,0,0,.4),0 8px 32px rgba(0,0,0,.25)}
.tbl thead th{
  background:linear-gradient(180deg,var(--surface-3) 0%,var(--surface-2) 100%);
  color:var(--muted-2);font-size:9px;text-transform:uppercase;
  letter-spacing:.08em;font-weight:700;text-align:right;
  padding:10px 13px;border-bottom:1px solid var(--border-2);
  white-space:nowrap;position:sticky;top:0;z-index:2;
  backdrop-filter:blur(8px)}
.tbl thead th:first-child{text-align:left}
.tbl thead th.now-hdr{color:var(--text);font-weight:800;letter-spacing:.04em}
.tbl tbody td{padding:5px 13px;border-bottom:1px solid var(--border);
  text-align:right;vertical-align:middle;
  font-variant-numeric:tabular-nums;white-space:nowrap;
  transition:background .08s}
.tbl tbody td:first-child{text-align:left}
.tbl tbody tr:last-child td{border-bottom:none}
.tbl tbody tr:hover td{background:rgba(255,255,255,.015)}

/* ── Section divider rows ── */
.section-divider td{background:var(--bg);color:var(--muted-2);font-size:8.5px;
  text-transform:uppercase;letter-spacing:.1em;font-weight:700;
  padding:14px 13px 4px;border-bottom:1px solid var(--border)}

/* ── Position header rows ── */
.pos-header td{background:var(--surface-2);padding:8px 13px;
  border-top:1px solid var(--border);border-bottom:1px solid rgba(255,255,255,.03)}
.pos-header td:first-child{border-left:3px solid var(--border-2)}
.pos-header.hr-green td:first-child{border-left-color:var(--green)}
.pos-header.hr-amber td:first-child{border-left-color:var(--amber)}
.pos-header.hr-red   td:first-child{border-left-color:var(--red)}
.pos-header.hr-none  td:first-child{border-left-color:var(--muted-2)}
.pos-header .protocol{font-weight:600;font-size:13px}
.pos-header .chain{display:inline-block;font-size:8px;padding:1px 6px;border-radius:999px;
  background:var(--surface-3);color:var(--muted-2);border:1px solid var(--border);
  margin-left:7px;vertical-align:2px;text-transform:uppercase;letter-spacing:.05em}
.pos-header .comp{font-size:10px;color:var(--muted);margin-left:8px;
  font-family:ui-monospace,Menlo,monospace}
.pos-header .hr-pill{margin-left:8px;vertical-align:1px}

/* ── Sub-rows (Collateral / Debt / Net / Balance) ── */
.subrow td{padding:4px 13px}
.subrow td:first-child{padding-left:28px;font-size:9px;font-weight:700;
  text-transform:uppercase;letter-spacing:.09em}
.subrow.collat td:first-child{color:var(--supply)}
.subrow.debt   td:first-child{color:var(--borrow)}
.subrow.net    td:first-child,.subrow.bal td:first-child{color:var(--muted)}
.subrow.net td.now-cell,.subrow.bal td.now-cell{font-weight:800;font-size:14px}
.subrow:last-of-type td{padding-bottom:10px}

/* ── Cell value classes ── */
.now-cell{font-weight:700;color:var(--text)}
.now-cell.neg{color:var(--red)}
.cell-up{color:var(--green);font-weight:600}
.cell-down{color:var(--red);font-weight:600}
.cell-flat{color:rgba(255,255,255,0.45)}
.cell-empty{color:rgba(255,255,255,0.18);font-size:11px}
.cell-new{color:var(--amber);font-size:9.5px;font-weight:700;letter-spacing:.04em}

/* ── Wallet rows ── */
.wallet-row td{padding:6px 13px}
.wallet-row td:first-child{padding-left:14px}
.wallet-row .tok-name{font-weight:700;font-size:12.5px}
.wallet-row .tok-chain{display:inline-block;font-size:8px;padding:1px 5px;border-radius:999px;
  background:var(--surface-2);color:var(--muted-2);border:1px solid var(--border);
  margin-left:6px;vertical-align:1px;text-transform:uppercase;letter-spacing:.04em}
.wallet-row .tok-qty{font-family:ui-monospace,Menlo,monospace;font-size:10px;
  color:var(--muted-2);margin-left:6px}

/* ── Totals row ── */
.totals-row td{
  background:linear-gradient(90deg,rgba(129,140,248,.08),rgba(129,140,248,.03));
  font-weight:800;font-size:14px;letter-spacing:-.01em;
  border-top:2px solid var(--border-2);padding:10px 13px}
.totals-row td:first-child{padding-left:14px}

/* ── Closed positions table ── */
.cls-tbl{width:100%;border-collapse:separate;border-spacing:0;
  background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);overflow:hidden}
.cls-tbl thead th{background:var(--surface-2);color:var(--muted-2);font-size:9px;
  text-transform:uppercase;letter-spacing:.08em;font-weight:700;text-align:right;
  padding:9px 13px;border-bottom:1px solid var(--border)}
.cls-tbl thead th:first-child{text-align:left}
.cls-tbl tbody td{padding:7px 13px;border-bottom:1px solid var(--border);
  text-align:right;color:var(--muted);font-size:12px;font-variant-numeric:tabular-nums}
.cls-tbl tbody td:first-child{text-align:left;color:var(--text)}
.cls-tbl tbody tr:last-child td{border-bottom:none}

.footnote{color:var(--muted-2);font-size:9.5px;margin-top:28px;text-align:center;line-height:2}
"""



def _lh_label(h: int) -> str:
    if h == 168: return "7d"
    if h == 336: return "14d"
    return f"{h}h"


EXTRA_CSS = """
/* ── Page nav ── */
.page-nav{display:flex;gap:4px;margin-bottom:20px}
.page-nav a{font-size:10.5px;font-weight:700;padding:5px 16px;border-radius:8px;
  border:1px solid var(--border);color:var(--muted);text-decoration:none;
  background:var(--surface);letter-spacing:.03em;transition:all .15s}
.page-nav a:hover{color:var(--text);border-color:var(--border-2);background:var(--surface-2)}
.page-nav a.active{color:var(--indigo);background:var(--indigo-dim);
  border-color:rgba(129,140,248,.3)}

/* ── Summary strip — Net hero + Collateral/Debt below ── */
.summary{display:grid;grid-template-columns:1fr 1fr;
  grid-template-rows:auto auto;gap:10px;margin-bottom:16px}
.summary-net{grid-column:1/-1}
.cell{background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);padding:14px 18px;
  transition:border-color .15s}
.cell:hover{border-color:var(--border-2)}
.cell-net{background:linear-gradient(135deg,rgba(129,140,248,.06) 0%,var(--surface) 80%);
  border-color:rgba(129,140,248,.25);padding:18px 22px}
.cell-collat{border-left:3px solid var(--supply)}
.cell-debt{border-left:3px solid var(--borrow)}
.cell .label{font-size:8px;color:var(--muted-2);text-transform:uppercase;
  letter-spacing:.1em;font-weight:700;margin-bottom:5px}
.cell-net .label{font-size:9px;color:var(--indigo);letter-spacing:.08em}
.cell .value{font-size:20px;font-weight:900;letter-spacing:-.04em;
  font-variant-numeric:tabular-nums;color:var(--text)}
.cell-net .value{font-size:32px;letter-spacing:-.05em;color:var(--text)}
.cell-collat .value{font-size:18px;color:var(--supply)}
.cell-debt .value,.cell-val-debt{font-size:18px;color:var(--borrow)!important}

/* ── Benchmark strip ── */
.bench-strip{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:22px}
.bench-card{background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);padding:15px 18px 13px;
  min-width:130px;flex:1;
  transition:border-color .15s,transform .1s;
  display:flex;flex-direction:column;gap:0}
.bench-card:hover{border-color:var(--border-2);transform:translateY(-1px)}
.bench-card-day0{
  background:linear-gradient(135deg,rgba(129,140,248,.07) 0%,var(--surface) 70%);
  border-color:rgba(129,140,248,.3)}
.bench-lbl{font-size:8px;color:var(--muted-2);text-transform:uppercase;
  letter-spacing:.1em;font-weight:700;margin-bottom:8px}
.bench-delta-up{font-size:26px;font-weight:900;font-variant-numeric:tabular-nums;
  letter-spacing:-.04em;line-height:1;color:var(--green);
  text-shadow:0 0 20px rgba(45,212,160,.3)}
.bench-delta-dn{font-size:26px;font-weight:900;font-variant-numeric:tabular-nums;
  letter-spacing:-.04em;line-height:1;color:var(--red);
  text-shadow:0 0 20px rgba(244,112,112,.25)}
.bench-delta-flat{font-size:26px;font-weight:900;font-variant-numeric:tabular-nums;
  letter-spacing:-.04em;line-height:1;color:var(--muted-2)}
.bench-pct-up{font-size:12px;font-weight:700;color:var(--green);
  margin-top:3px;letter-spacing:-.01em;opacity:.85}
.bench-pct-dn{font-size:12px;font-weight:700;color:var(--red);
  margin-top:3px;letter-spacing:-.01em;opacity:.85}
.bench-pct-flat{font-size:12px;font-weight:600;color:var(--muted-2);margin-top:3px}
.bench-base{font-size:9.5px;color:var(--muted-2);
  font-variant-numeric:tabular-nums;margin-top:6px;border-top:1px solid var(--border);
  padding-top:6px}
.bench-na{color:var(--muted-2);font-size:18px;font-weight:900;line-height:1}

/* ── Token group headers ── */
.tok-group-hdr td{
  padding:16px 14px 5px;
  border-top:1px solid var(--border-2);
  background:transparent}
.tok-group-hdr:first-child td{border-top:none;padding-top:8px}
.tok-group-name{
  font-size:17px;font-weight:900;letter-spacing:-.04em;color:var(--text);
  display:block;line-height:1.2}
.tok-group-sub{font-size:10px;color:var(--muted-2);margin-top:2px;display:block}

/* ── Position header ── */
.pos-header td:first-child{padding-left:22px}
.proto-name{font-weight:600;font-size:12px;color:var(--text)}
.proto-chain{display:inline-block;font-size:8px;padding:1px 6px;border-radius:999px;
  background:var(--surface-3);color:var(--muted-2);border:1px solid var(--border);
  margin-left:6px;vertical-align:2px;text-transform:uppercase;letter-spacing:.05em}
.proto-pair{font-size:9px;color:var(--muted-2);margin-left:8px;
  font-family:ui-monospace,Menlo,monospace}
.hr-pill{margin-left:9px;vertical-align:1px}

/* ── Fund hero ── */
.fund-hero{
  display:flex;justify-content:space-between;align-items:flex-start;
  gap:24px;flex-wrap:wrap;
  background:linear-gradient(135deg,rgba(129,140,248,.06) 0%,rgba(45,212,160,.03) 100%);
  border:1px solid rgba(129,140,248,.2);
  border-radius:var(--radius);padding:28px 32px 24px;margin-bottom:24px}
.fund-hero-left{display:flex;flex-direction:column;gap:6px;min-width:200px}
.fund-badge{display:flex;align-items:center;gap:6px;font-size:9px;font-weight:700;
  text-transform:uppercase;letter-spacing:.12em;color:var(--muted-2);margin-bottom:2px}
.fund-name{font-size:13px;font-weight:700;letter-spacing:.02em;color:var(--muted);
  text-transform:uppercase}
.fund-balance{font-size:48px;font-weight:900;letter-spacing:-.06em;
  font-variant-numeric:tabular-nums;color:var(--text);line-height:1;margin:-2px 0 4px}
.fund-meta{font-size:9.5px;color:var(--muted-2);font-family:ui-monospace,Menlo,monospace;
  letter-spacing:.02em}

/* ── Performance pills ── */
.fund-hero-right{display:flex;gap:8px;flex-wrap:wrap;align-items:flex-start;
  padding-top:4px}
.perf-pill{background:var(--surface-2);border:1px solid var(--border-2);
  border-radius:10px;padding:11px 15px;min-width:100px;
  transition:border-color .15s,transform .1s}
.perf-pill:hover{transform:translateY(-1px);border-color:var(--border-2)}
.perf-pill.perf-up{border-color:rgba(45,212,160,.25);background:rgba(45,212,160,.05)}
.perf-pill.perf-dn{border-color:rgba(244,112,112,.22);background:rgba(244,112,112,.04)}
.perf-pill.perf-flat{border-color:var(--border)}
.perf-pill.perf-na{border-color:var(--border);opacity:.5}
.perf-lbl{font-size:8px;text-transform:uppercase;letter-spacing:.1em;font-weight:700;
  color:var(--muted-2);margin-bottom:5px}
.perf-val{font-size:20px;font-weight:900;font-variant-numeric:tabular-nums;
  letter-spacing:-.03em;line-height:1}
.perf-pill.perf-up .perf-val{color:var(--green)}
.perf-pill.perf-dn .perf-val{color:var(--red)}
.perf-pill.perf-flat .perf-val,.perf-pill.perf-na .perf-val{color:var(--muted-2)}
.perf-pct{font-size:10.5px;font-weight:700;margin-top:3px}
.perf-pill.perf-up .perf-pct{color:var(--green);opacity:.8}
.perf-pill.perf-dn .perf-pct{color:var(--red);opacity:.8}
.perf-pill.perf-flat .perf-pct{color:var(--muted-2)}

/* ── Leverage detail cards ── */
.lev-cards{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:10px}
.lev-card{background:var(--surface);border:1px solid var(--border);
  border-radius:var(--radius);padding:14px 18px}
.lev-collat{border-left:3px solid var(--supply)}
.lev-debt{border-left:3px solid var(--borrow)}
.lev-net2{border-left:3px solid var(--indigo)}
.lev-lbl{font-size:8px;text-transform:uppercase;letter-spacing:.1em;font-weight:700;
  color:var(--muted-2);margin-bottom:5px}
.lev-val{font-size:20px;font-weight:900;font-variant-numeric:tabular-nums;
  letter-spacing:-.04em;color:var(--text)}
.lev-val-debt{color:var(--borrow)!important}
"""


def render_dashboard_html(enriched: dict, generated_at: str) -> str:
    """Token-first positions page with benchmark strip and per-position lookback table."""
    import json as _json
    payload_json = _json.dumps(enriched, separators=(",", ":"), default=str)
    payload_safe = payload_json.replace("</", "<\\/")
    # Column headers — compact labels
    lh_hdrs = "".join(f'<th>{_lh_label(h)}</th>' for h in LOOKBACKS_H)

    js = """
(function(){
  var D=JSON.parse(document.getElementById('payload').textContent);
  var L=D.lookbacksHours||[12,24,36,48,168,336];

  function fmtUsd(n,dec){
    if(n==null)return'—';
    return new Intl.NumberFormat('en-US',{style:'currency',currency:'USD',
      minimumFractionDigits:dec||0,maximumFractionDigits:dec||0}).format(n);
  }
  function fmtPct(p){if(p==null)return'—';return(p>=0?'+':'')+p.toFixed(2)+'%';}
  function hrCls(h){return h==null?'gray':(h<1.05?'red':(h<1.15?'amber':'green'));}

  // ── Time-series cell colouring ───────────────────────────────────────────────
  // Each cell shows its VALUE coloured relative to the NEXT-OLDER cell.
  // The oldest populated cell is always flat (it's the baseline).
  // The "Now" cell is coloured relative to the 12h cell (most recent history).
  // Flat $20 threshold — consistent across all position sizes.
  function thr(v){return 20;}
  function dirCls(val,prevVal,kind){
    if(prevVal==null)return'cell-flat';
    var diff=val-prevVal;
    if(Math.abs(diff)<thr(val))return'cell-flat';
    var up=kind==='liability'?(diff<0):(diff>0);
    return up?'cell-up':'cell-down';
  }
  function newCell(){return'<td class="cell-new">new</td>';}
  function emptyCell(){return'<td class="cell-empty">—</td>';}
  function hrHdrCls(hr){return hr==null?'hr-none':(hr<1.05?'hr-red':(hr<1.15?'hr-amber':'hr-green'));}

  // Build a full data row with time-series coloring.
  // hist = p.history dict, cur = current value, getH = fn(hh)->number
  function mRow(hist,lbl,kl,cur,kind,getH){
    // Collect historical values in L order (12h=index0 is most recent, 14d=last is oldest)
    var hvals=L.map(function(h){
      var hh=hist&&hist[String(h)];
      if(!hh||hh.absent)return null;
      return getH(hh);
    });
    // Find the rightmost (oldest) non-null index — it gets flat styling
    var oldestIdx=-1;
    for(var i=hvals.length-1;i>=0;i--){if(hvals[i]!=null){oldestIdx=i;break;}}

    // Now cell: compare vs hvals[0] (12h), or flat if no history
    var nowCls='now-cell'+(cur<0?' neg':'')+' '+dirCls(cur,hvals[0],kind);
    var nowTd='<td class="'+nowCls.trim()+'">'+fmtUsd(cur)+'</td>';

    // History cells: each compares to its immediate right neighbour (next older)
    var hc=hvals.map(function(val,i){
      var hh=hist&&hist[String(L[i])];
      if(val==null){if(!hh)return emptyCell();return hh.absent?newCell():emptyCell();}
      if(i===oldestIdx)return'<td class="cell-flat">'+fmtUsd(val)+'</td>';
      // Find next older non-null
      var prevVal=null;
      for(var j=i+1;j<hvals.length;j++){if(hvals[j]!=null){prevVal=hvals[j];break;}}
      return'<td class="'+dirCls(val,prevVal,kind)+'">'+fmtUsd(val)+'</td>';
    }).join('');

    return'<tr class="subrow '+kl+'"><td>'+lbl+'</td>'+nowTd+hc+'</tr>';
  }

  // ── Fund hero ────────────────────────────────────────────────────────────────
  var t=D.totals,wt=t.walletTokensTotal||0,eq=D.reportedTotal||(t.net+wt);
  var netPos=(t.net||0)+(wt||0);

  var balEl=document.getElementById('fund-balance');
  if(balEl)balEl.textContent=fmtUsd(netPos);
  var metaEl=document.getElementById('fund-meta');
  if(metaEl)metaEl.textContent='as of '+D.capturedAt+' UTC  \u00b7  '+D.wallet.slice(0,6)+'\u2026'+D.wallet.slice(-4);
  var hnEl=document.getElementById('history-note');
  if(hnEl)hnEl.textContent=D.snapshotCount+' snapshots \u00b7 Day 0: '+((D.day0&&D.day0.at)||'\u2014');

  // Performance pills: Day 0 + benchmark periods
  function perfPill(label,delta,pct){
    if(delta==null)return'<div class="perf-pill perf-na"><div class="perf-lbl">'+label+'</div><div class="perf-val">no data</div></div>';
    var flat=pct!=null&&Math.abs(pct)<0.05;
    var up=delta>=0;
    var cls=flat?'perf-flat':(up?'perf-up':'perf-dn');
    var sign=(up&&!flat)?'+':'';
    return'<div class="perf-pill '+cls+'">'+
      '<div class="perf-lbl">'+label+'</div>'+
      '<div class="perf-val">'+(flat?'\u2014':sign+fmtUsd(Math.abs(delta),0))+'</div>'+
      '<div class="perf-pct">'+(flat?'flat':((up?'+':'')+pct.toFixed(2)+'%'))+'</div>'+
      '</div>';
  }
  var ph='';
  if(D.day0&&D.day0.delta!=null)ph+=perfPill('Since day 0',D.day0.delta,D.day0.pct);
  (D.benchmarks||[]).forEach(function(b){ph+=perfPill(b.label,b.delta,b.pct);});
  var perfEl=document.getElementById('fund-perf');
  if(perfEl)perfEl.innerHTML=ph;

  // ── Leverage detail (bottom of page) ─────────────────────────────────────────
  var levEl=document.getElementById('leverage-section');
  if(levEl){
    levEl.innerHTML=
      '<div class="section-row" style="margin-top:32px">'+
        '<div class="section-title">Leverage detail</div>'+
        '<div class="legend" style="font-size:9px">Gross collateral &amp; debt across all positions</div>'+
      '</div>'+
      '<div class="lev-cards">'+
        '<div class="lev-card lev-collat"><div class="lev-lbl">Total collateral</div><div class="lev-val">'+fmtUsd(t.sup)+'</div></div>'+
        '<div class="lev-card lev-debt"><div class="lev-lbl">Total debt</div><div class="lev-val lev-val-debt">'+fmtUsd(t.bor)+'</div></div>'+
        '<div class="lev-card lev-net2"><div class="lev-lbl">DeFi net</div><div class="lev-val">'+fmtUsd(t.net)+'</div></div>'+
      '</div>';
  }


  // ── Token-first grouped table ────────────────────────────────────────────────
  var COLSPAN=2+L.length;

  // Group DeFi positions by their primary token(s)
  // Key = sorted supply tokens joined (e.g. "FRAX+USDC" or "siUSD")
  function tokKey(p){
    return(p.supplied||[]).map(function(x){return x.token;}).sort().join('+') || 'other';
  }
  var groups={},groupOrder=[];
  D.positions.slice().sort(function(a,b){return b.net-a.net;}).forEach(function(p){
    var k=tokKey(p);
    if(!groups[k]){groups[k]=[];groupOrder.push(k);}
    groups[k].push(p);
  });
  // Sort groups by total net desc
  groupOrder.sort(function(a,b){
    var sa=groups[a].reduce(function(s,p){return s+p.net;},0);
    var sb=groups[b].reduce(function(s,p){return s+p.net;},0);
    return sb-sa;
  });

  var rows=[];
  var totalEquity=netPos;  // same as fund hero — netPos = t.net + wallet tokens

  if(groupOrder.length){
    rows.push('<tr class="section-divider"><td colspan="'+COLSPAN+'">DeFi positions</td></tr>');
  }
  groupOrder.forEach(function(k){
    var ps=groups[k];
    var groupNet=ps.reduce(function(s,p){return s+p.net;},0);
    var groupSup=ps.reduce(function(s,p){return s+p.supSum;},0);
    var groupBor=ps.reduce(function(s,p){return s+p.borSum;},0);
    var tokDisplay=k.replace(/\+/g,' + ');
    var protos=ps.map(function(p){return p.protocol;}).filter(function(v,i,a){return a.indexOf(v)===i;}).join(', ');
    var chains=ps.map(function(p){return p.chain;}).filter(function(v,i,a){return a.indexOf(v)===i;}).join(', ');
    var groupSubtitle=protos+(chains?' · '+chains:'');

    // Group header row — shows token name + aggregated net
    rows.push(
      '<tr class="tok-group-hdr">'+
        '<td colspan="'+COLSPAN+'">'+
          '<span class="tok-group-name">'+tokDisplay+'</span>'+
          (ps.length>1?'<span class="tok-group-sub">'+groupSubtitle+'</span>':'')+
        '</td>'+
      '</tr>'
    );

    // Individual position rows within the group
    ps.forEach(function(p){
      var bor=p.borrowed&&p.borrowed.length?p.borrowed.map(function(x){return x.token;}).join('+'):null;
      var hrLbl=p.healthPct!=null?p.healthPct+'% to liq':(p.healthRate!=null?p.healthRate.toFixed(4):null);
      var hrCell=hrLbl!=null?'<span class="pill '+hrCls(p.healthRate)+' hr-pill">'+hrLbl+'</span>':'<span class="pill gray hr-pill">no debt</span>';
      // Protocol + chain in smaller text; borrow pair if present
      var subInfo='<span class="proto-name">'+p.protocol+'</span>'+
        '<span class="proto-chain">'+p.chain+'</span>'+
        (bor?'<span class="proto-pair">borrows '+bor+'</span>':'')+
        hrCell;
      rows.push('<tr class="pos-header '+hrHdrCls(p.healthRate)+'"><td colspan="'+COLSPAN+'">'+subInfo+'</td></tr>');
      if(p.borSum&&p.borSum>0.5){
        rows.push(mRow(p.history,'Collateral','collat',p.supSum,'asset',function(h){return h.supSum;}));
        rows.push(mRow(p.history,'Debt','debt',p.borSum,'liability',function(h){return h.borSum;}));
        rows.push(mRow(p.history,'Net','net',p.net,'asset',function(h){return h.net;}));
      }else{
        rows.push(mRow(p.history,'Balance','bal',p.supSum,'asset',function(h){return h.supSum;}));
      }
    });
  });

  // Wallet tokens
  var sW=(D.walletTokens||[]).slice().sort(function(a,b){return b.usd-a.usd;});
  if(sW.length){rows.push('<tr class="section-divider"><td colspan="'+COLSPAN+'">Wallet tokens</td></tr>');}
  sW.forEach(function(w){
    var amt=w.amount>=1?Math.round(w.amount).toLocaleString('en-US'):w.amount.toFixed(4);
    var wvals=L.map(function(h){var hh=w.history&&w.history[String(h)];return(hh&&!hh.absent)?hh.usd:null;});
    var wOldest=-1;for(var i=wvals.length-1;i>=0;i--){if(wvals[i]!=null){wOldest=i;break;}}
    var wNowCls='now-cell '+dirCls(w.usd,wvals[0],'asset');
    var whc=wvals.map(function(val,i){
      var hh=w.history&&w.history[String(L[i])];
      if(val==null){if(!hh)return emptyCell();return hh.absent?newCell():emptyCell();}
      if(i===wOldest)return'<td class="cell-flat">'+fmtUsd(val)+'</td>';
      var pv=null;for(var j=i+1;j<wvals.length;j++){if(wvals[j]!=null){pv=wvals[j];break;}}
      return'<td class="'+dirCls(val,pv,'asset')+'">'+fmtUsd(val)+'</td>';
    }).join('');
    rows.push('<tr class="wallet-row"><td><span class="tok-name">'+w.token+'</span>'+
      '<span class="tok-chain">'+w.chain+'</span><span class="tok-qty">'+amt+'</span></td>'+
      '<td class="'+wNowCls.trim()+'">'+fmtUsd(w.usd)+'</td>'+whc+'</tr>');
  });

  // Totals row — time-series coloring same logic
  var tvals=L.map(function(h){
    var s=0,any=false;
    D.positions.forEach(function(p){var hh=p.history&&p.history[String(h)];if(hh&&!hh.absent&&hh.net!=null){s+=hh.net;any=true;}});
    sW.forEach(function(w){var hh=w.history&&w.history[String(h)];if(hh&&!hh.absent&&hh.usd!=null){s+=hh.usd;any=true;}});
    return any?s:null;
  });
  var tOldest=-1;for(var ti=tvals.length-1;ti>=0;ti--){if(tvals[ti]!=null){tOldest=ti;break;}}
  var tNowCls='now-cell '+dirCls(totalEquity,tvals[0],'asset');
  var tc=tvals.map(function(val,i){
    if(val==null)return emptyCell();
    if(i===tOldest)return'<td class="cell-flat">'+fmtUsd(val)+'</td>';
    var pv=null;for(var j=i+1;j<tvals.length;j++){if(tvals[j]!=null){pv=tvals[j];break;}}
    return'<td class="'+dirCls(val,pv,'asset')+'">'+fmtUsd(val)+'</td>';
  }).join('');
  rows.push('<tr class="totals-row"><td>Total equity</td><td class="'+tNowCls.trim()+'">'+fmtUsd(totalEquity)+'</td>'+tc+'</tr>');
  document.getElementById('tbody').innerHTML=rows.join('');

  // Closed positions
  if(D.closedPositions&&D.closedPositions.length){
    document.getElementById('closed').innerHTML=
      '<div class="section-row" style="margin-top:26px"><div class="section-title">Closed since first snapshot</div></div>'+
      '<table class="cls-tbl"><thead><tr><th style="text-align:left">Position</th>'+
      '<th>Last seen</th><th>Collateral</th><th>Debt</th><th>Net</th></tr></thead><tbody>'+
      D.closedPositions.map(function(q){
        return'<tr><td><b>'+q.protocol+'</b><span class="tok-chain">'+q.chain+
          '</span> <span style="color:var(--muted);font-size:11px">'+q.name+'</span></td>'+
          '<td style="color:var(--muted-2);font-size:11px">'+q.lastSeenAt+'</td>'+
          '<td>'+fmtUsd(q.supSum)+'</td><td>'+fmtUsd(q.borSum)+'</td><td>'+fmtUsd(q.net)+'</td></tr>';
      }).join('')+'</tbody></table>';
  }

  // Live EVM total
  fetch('https://api.rabby.io/v1/user/total_balance?id='+D.wallet,{headers:{'accept':'application/json'}})
  .then(function(r){return r.json();}).then(function(j){
    var live=j&&j.total_usd_value;if(live==null)return;
    var diff=live-totalEquity,pct=totalEquity?(diff/totalEquity*100):0;
    var cls=Math.abs(pct)<0.01?'cell-flat':(diff>=0?'cell-up':'cell-down');
    document.getElementById('live-line').innerHTML=
      'Live EVM: <strong>'+fmtUsd(live)+'</strong> <span class="'+cls+'" style="font-size:10px">'+
      (diff>=0?'▲':'▼')+' '+Math.abs(pct).toFixed(2)+'%</span>';
  }).catch(function(){});
})();
"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Fund Dashboard</title>
<style>{DASHBOARD_CSS}{EXTRA_CSS}</style>
</head>
<body>
<nav class="page-nav">
  <a href="index.html" class="active">Overview</a>
  <a href="history.html">History</a>
</nav>

<!-- ── Fund hero ─────────────────────────────────────────────────── -->
<div class="fund-hero">
  <div class="fund-hero-left">
    <div class="fund-badge"><span class="live-dot"></span>Live</div>
    <div class="fund-name">Portfolio</div>
    <div class="fund-balance" id="fund-balance">—</div>
    <div class="fund-meta" id="fund-meta"></div>
  </div>
  <div class="fund-hero-right" id="fund-perf"></div>
</div>

<!-- ── Holdings table ────────────────────────────────────────────── -->
<div class="section-row" style="margin-top:20px">
  <div class="section-title">Holdings</div>
  <div class="legend">
    <span><span class="swatch" style="background:#2dd4a0"></span>Up vs prior period</span>
    <span><span class="swatch" style="background:#f47070"></span>Down vs prior period</span>
  </div>
</div>
<table class="tbl">
  <thead><tr><th>Token / Position</th><th class="now-hdr">Balance</th>{lh_hdrs}</tr></thead>
  <tbody id="tbody"></tbody>
</table>
<div id="closed"></div>

<!-- ── Leverage detail (bottom) ─────────────────────────────────── -->
<div class="leverage-section" id="leverage-section"></div>

<div class="footnote">Updated {generated_at} UTC &middot; Auto-refreshes every 6h &middot; Positions &gt;$50 &middot; <span id="history-note"></span></div>
<script id="payload" type="application/json">{payload_safe}</script>
<script>{js}</script>
</body></html>"""
    return html


def render_history_html(snapshots: list[dict], generated_at: str) -> str:
    """Full chronological snapshot history page."""
    sorted_snaps = sorted(snapshots, key=lambda s: s.get("capturedAt", ""), reverse=True)
    d0_at = ""
    d0_path = Path(__file__).parent.parent / "outputs/snapshots/day0.json"
    if d0_path.exists():
        try:
            d0_at = json.loads(d0_path.read_text()).get("capturedAt", "")
        except Exception:
            pass
    rows_html = ""
    for i, s in enumerate(sorted_snaps):
        eq   = s.get("reportedTotal") or 0
        prev = sorted_snaps[i + 1].get("reportedTotal") or 0 if i + 1 < len(sorted_snaps) else None
        delta = round(eq - prev, 0) if prev is not None else None
        pct   = round((eq - prev) / prev * 100, 2) if prev else None
        count = (s.get("totals") or {}).get("count", 0)
        lr    = (s.get("totals") or {}).get("lowest_hr")
        tag   = ' <span style="font-size:10px;color:var(--amber)">Day 0</span>' if s.get("capturedAt") == d0_at else ""
        if delta is not None:
            cls   = "cell-up" if delta >= 0 else "cell-down"
            sign  = "+" if delta >= 0 else ""
            d_html = f'<span class="{cls}">{sign}${abs(delta):,.0f}</span>'
            if pct is not None:
                d_html += f' <span class="{cls}" style="font-size:10px">({sign}{pct:.2f}%)</span>'
        else:
            d_html = "—"
        hr_cls  = "" if lr is None else ("cell-down" if lr < 1.05 else ("cell-flat" if lr < 1.15 else "cell-up"))
        hr_html = f'<span class="{hr_cls}">{lr:.4f}</span>' if lr else "—"
        rows_html += f"""<tr>
          <td style="text-align:left">{s.get("capturedAt","")}{tag}</td>
          <td>${eq:,.0f}</td><td>{d_html}</td>
          <td>{count}</td><td>{hr_html}</td>
        </tr>\n"""
    nav_css = """
.page-nav{display:flex;gap:4px;margin-bottom:16px}
.page-nav a{font-size:11px;font-weight:600;padding:5px 14px;border-radius:6px;
  border:1px solid var(--border);color:var(--muted);text-decoration:none;background:var(--surface)}
.page-nav a:hover{color:var(--text);border-color:var(--border-2)}
.page-nav a.active{color:var(--text);background:var(--surface-2);border-color:var(--border-2)}
"""
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>DeFi History</title>
<style>{DASHBOARD_CSS}{nav_css}</style>
</head>
<body>
<nav class="page-nav">
  <a href="index.html">Positions</a>
  <a href="history.html" class="active">History</a>
</nav>
<header class="page">
  <div class="hdr-left">
    <div class="hdr-title"><span class="live-dot"></span><h1>Snapshot History</h1></div>
  </div>
  <div class="hdr-right">
    <div class="snap-time">{len(sorted_snaps)} snapshots total</div>
    <div class="source">12h refresh cadence</div>
  </div>
</header>
<table class="tbl" style="margin-top:12px">
  <thead><tr>
    <th style="text-align:left">Captured (UTC)</th>
    <th>Total equity</th><th>vs previous</th>
    <th>Positions</th><th>Lowest HR</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="footnote">Generated {generated_at} UTC</div>
</body></html>"""
    return html


def render_and_write_dashboard(snapshots_dir: Path, dashboard_path: Path, current: dict):
    snapshots = load_all_snapshots(snapshots_dir)
    enriched  = build_history(current, snapshots)
    gen_at    = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    html = render_dashboard_html(enriched, gen_at)
    dashboard_path.parent.mkdir(parents=True, exist_ok=True)
    dashboard_path.write_text(html, encoding="utf-8")
    print(f"[refresh] Positions page ({len(html):,} bytes, {len(snapshots)} snapshots) → {dashboard_path}")

    history_path = dashboard_path.parent / "history.html"
    hist_html    = render_history_html(snapshots + [current], gen_at)
    history_path.write_text(hist_html, encoding="utf-8")
    print(f"[refresh] History page ({len(hist_html):,} bytes) → {history_path}")


# ────────────────────────── Solana wallet ──────────────────────────

def _solana_rpc(method: str, params: list):
    payload = json.dumps({"jsonrpc": "2.0", "id": 1, "method": method, "params": params}).encode()
    req = urllib.request.Request(SOL_RPC, data=payload,
                                  headers={"Content-Type": "application/json", "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=25) as r:
        return json.loads(r.read()).get("result")


def solana_wallet_tokens(min_usd: float = FILTER_USD) -> list[dict]:
    """Fetch Solana wallet token balances via public RPC + Jupiter price API."""
    mints: dict[str, float] = {}

    for prog in [TOKEN_PROGRAM, TOKEN_2022_PROGRAM]:
        try:
            result = _solana_rpc("getTokenAccountsByOwner",
                                 [SOL_WALLET, {"programId": prog}, {"encoding": "jsonParsed"}])
            for acc in (result or {}).get("value", []):
                info = (acc.get("account", {}).get("data", {})
                           .get("parsed", {}).get("info", {}))
                mint = info.get("mint", "")
                amt = float((info.get("tokenAmount") or {}).get("uiAmount") or 0)
                if mint and amt > 0:
                    mints[mint] = mints.get(mint, 0) + amt
        except Exception:
            pass

    try:
        result = _solana_rpc("getBalance", [SOL_WALLET])
        lamports = result.get("value", 0) if isinstance(result, dict) else (result or 0)
        if lamports > 0:
            mints[SOL_MINT] = mints.get(SOL_MINT, 0) + lamports / 1e9
    except Exception:
        pass

    if not mints:
        return []

    try:
        ids = ",".join(mints.keys())
        req = urllib.request.Request(f"https://api.jup.ag/price/v2?ids={ids}",
                                      headers={"accept": "application/json", "User-Agent": UA})
        with urllib.request.urlopen(req, timeout=20) as r:
            price_data = json.loads(r.read()).get("data", {})
    except Exception:
        price_data = {}

    try:
        req = urllib.request.Request("https://tokens.jup.ag/tokens?tags=verified",
                                      headers={"accept": "application/json", "User-Agent": UA})
        with urllib.request.urlopen(req, timeout=20) as r:
            meta = {t["address"]: t for t in json.loads(r.read())}
    except Exception:
        meta = {}

    out = []
    for mint, amount in mints.items():
        price_info = price_data.get(mint, {})
        price = float(price_info.get("price") or 0)
        usd = amount * price
        if usd < min_usd:
            continue
        tok_meta = meta.get(mint, {})
        symbol = (tok_meta.get("symbol") or price_info.get("mintSymbol") or mint[:8]).upper()
        out.append({
            "chain": "Solana",
            "chainId": "sol",
            "token": symbol,
            "id": mint.lower(),
            "amount": amount,
            "usd": usd,
        })

    out.sort(key=lambda x: -x["usd"])
    return out


# ────────────────── Jupiter / Solana DeFi ──────────────────

def _parse_usd(s: str) -> float:
    return float(s.replace(",", "").replace("$", ""))


def _parse_jupiter_text(text: str) -> tuple[list[dict], float]:
    """Parse Jupiter portfolio page text → (positions, sol_net_worth)."""
    import re

    sol_net = 0.0
    m = re.search(r'Net Worth\$([\d,]+\.?\d*)', text)
    if m:
        sol_net = _parse_usd(m.group(1))

    positions = []

    # Each Loopscale lending block:
    # "LendingHealth 18%$52,676.82Supplied$201,328.94...Borrowed$148,652.13..."
    lending_pat = re.compile(
        r'LendingHealth\s+(\d+)%\$([\d,]+\.?\d*)'
        r'Supplied\$([\d,]+\.?\d*)'
        r'TokenBalance.*?Value(.*?)'
        r'Borrowed\$([\d,]+\.?\d*)'
        r'TokenBalance.*?Value(.*?)'
        r'(?=LendingHealth|\Z|Kamino)',
        re.DOTALL,
    )
    tok_pat = re.compile(
        r'([A-Za-z][A-Za-z0-9]+)\s+([\d,]+(?:\.\d+)?)\s+\S+\$([\d.]+)[^$]*\$([\d,]+\.?\d*)'
    )

    for bm in lending_pat.finditer(text):
        health_pct = int(bm.group(1))
        net_usd    = _parse_usd(bm.group(2))
        sup_usd    = _parse_usd(bm.group(3))
        sup_block  = bm.group(4)
        bor_usd    = _parse_usd(bm.group(5))
        bor_block  = bm.group(6)

        def parse_toks(block):
            out = []
            for tm in tok_pat.finditer(block):
                sym, amt_s, _, usd_s = tm.groups()
                out.append({"token": sym, "amount": float(amt_s.replace(",", "")),
                            "usd": _parse_usd(usd_s)})
            return out

        sup_toks = parse_toks(sup_block) or [{"token": "?", "amount": 0.0, "usd": sup_usd}]
        bor_toks = parse_toks(bor_block) or [{"token": "?", "amount": 0.0, "usd": bor_usd}]

        positions.append({
            "protocol": "Loopscale",
            "chain": "Solana",
            "chainId": "sol",
            "type": "Lending",
            "rabbyType": "Lending",
            "healthRate": round(1.0 + health_pct / 100, 4),
            "healthPct": health_pct,
            "supplied": sup_toks,
            "borrowed": bor_toks,
            "rewards": [],
            "description": f"Health {health_pct}%",
            "_supSum": sup_usd,
            "_borSum": bor_usd,
            "_rewSum": 0.0,
            "_net": net_usd,
            "_poolId": "",
            "_source": "jupiter",
        })

    # Kamino rewards
    km = re.search(r'Kamino\$([\d,]+\.?\d*)Farming\$([\d,]+\.?\d*)Rewards', text)
    if km:
        rew_usd = _parse_usd(km.group(2))
        if rew_usd >= 0.01:
            rew_toks = []
            after = text[km.end():]
            for tm in re.finditer(r'\b([A-Z][A-Z0-9]{1,7})\s+Claimable\s*([\d,]+\.?\d*)\s+\S+\$([\d,]+\.?\d*)', after):
                sym, amt_s, usd_s = tm.groups()
                rew_toks.append({"token": sym, "amount": float(amt_s.replace(",", "")),
                                 "usd": _parse_usd(usd_s)})
            positions.append({
                "protocol": "Kamino",
                "chain": "Solana",
                "chainId": "sol",
                "type": "Yield",
                "rabbyType": "Farming",
                "healthRate": None,
                "healthPct": None,
                "supplied": [],
                "borrowed": [],
                "rewards": rew_toks or [{"token": "USDC", "amount": 0.0, "usd": rew_usd}],
                "description": "Claimable rewards",
                "_supSum": 0.0,
                "_borSum": 0.0,
                "_rewSum": rew_usd,
                "_net": rew_usd,
                "_poolId": "",
                "_source": "jupiter",
            })

    return positions, sol_net


def _jup_token_meta() -> dict:
    """Fetch Jupiter verified token list → {mint_address: token_dict}."""
    try:
        req = urllib.request.Request(
            "https://tokens.jup.ag/tokens?tags=verified",
            headers={"accept": "application/json", "User-Agent": UA},
        )
        with urllib.request.urlopen(req, timeout=25) as r:
            return {t["address"]: t for t in json.loads(r.read())}
    except Exception as e:
        print(f"[refresh] Jupiter token meta failed: {e}", file=sys.stderr)
        return {}


def _jup_prices(mints: set) -> dict:
    """Fetch prices from Jupiter Price API v2 → {mint: {price: ...}}."""
    if not mints:
        return {}
    try:
        ids = ",".join(mints)
        req = urllib.request.Request(
            f"https://api.jup.ag/price/v2?ids={ids}",
            headers={"accept": "application/json", "User-Agent": UA},
        )
        with urllib.request.urlopen(req, timeout=25) as r:
            return json.loads(r.read()).get("data", {})
    except Exception as e:
        print(f"[refresh] Jupiter price API failed: {e}", file=sys.stderr)
        return {}


def _solana_mint_decimals(mint: str) -> int:
    """Get token decimals directly from Solana RPC for a mint address."""
    try:
        result = _solana_rpc("getAccountInfo", [mint, {"encoding": "jsonParsed"}])
        if result:
            info = (result.get("value") or {}).get("data", {}).get("parsed", {}).get("info", {})
            return int(info.get("decimals", 6))
    except Exception:
        pass
    return 6  # safe default (USDC/USDT)


USDC_MINT = "EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v"

# Hard-coded overrides for tokens not in Jupiter's verified list
KNOWN_SOL_MINTS: dict[str, dict] = {
    "EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v": {"symbol": "USDC",  "decimals": 6},
    "Es9vMFrzaCERmJfrF4H2FYD4KCoNkY11McCe8BenwNYB":  {"symbol": "USDT",  "decimals": 6},
    "So11111111111111111111111111111111111111112":    {"symbol": "SOL",   "decimals": 9},
    "5Y8NV33Vv7WbnLfq3zBcKSdYPrk7g2KoiQoe7M2tcxp5": {"symbol": "ONyc",  "decimals": 9},
    "3b8X44fLF9ooXaUm3hhSgjpmVs6rZZ3pPoGnGahc3Uu7": {"symbol": "PRIME", "decimals": 6},
}


def loopscale_positions(meta: dict | None = None) -> tuple[list[dict], float]:
    """Fetch active Loopscale lending positions via REST API — no browser required.

    API: POST https://tars.loopscale.com/v1/markets/loans/info
    Response shape: {"loanInfos": [{loan:{closed:bool,...}, ledgers:[...], collateral:[...]}]}
    Ratios are scaled ×1,000,000 (e.g. 900000 = 90% LTV).

    Returns (positions, total_net_usd).
    """
    url = "https://tars.loopscale.com/v1/markets/loans/info"
    payload = json.dumps({
        "borrowers": [SOL_WALLET],
        "page": 0,
        "pageSize": 50,
    }).encode()
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json", "accept": "application/json", "User-Agent": UA},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.loads(r.read())

    loan_infos = data.get("loanInfos", [])
    # Active = not closed AND has at least one active ledger
    active = [li for li in loan_infos
              if not li.get("loan", {}).get("closed", True) and li.get("ledgers")]
    print(f"[refresh] Loopscale API: {len(active)} active / {len(loan_infos)} total loans")
    if not active:
        return [], 0.0

    # Gather all mints we'll need for pricing
    mints: set[str] = set()
    for li in active:
        for lg in li.get("ledgers", []):
            mints.add(lg.get("principalMint") or "")
        for col in li.get("collateral", []):
            mints.add(col.get("assetMint") or "")
    mints.discard("")

    # Build event-based price fallback: latest usdPrice per mint from Loopscale events
    event_prices: dict[str, float] = {}
    for li in active:
        for ev in li.get("events", []):
            mint = ev.get("assetMint") or ""
            price = ev.get("usdPrice")
            if mint and price:
                event_prices[mint] = float(price)  # later events overwrite earlier ones

    if meta is None:
        meta = _jup_token_meta()
    price_data = _jup_prices(mints)

    positions = []
    total_net = 0.0

    for li in active:
        loan    = li.get("loan", {})
        loan_id = str(loan.get("address") or "")[:8]
        ledger  = li["ledgers"][0]  # one active ledger per loan
        cols    = li.get("collateral", [])
        col_entry = cols[0] if cols else {}

        prin_mint = ledger.get("principalMint") or ""
        col_mint  = col_entry.get("assetMint") or ""
        prin_raw  = int(ledger.get("principalDue") or 0)
        col_raw   = int(col_entry.get("amount") or 0)

        # Decimals: known-mints table → Jupiter meta → Solana RPC fallback
        prin_dec = int((KNOWN_SOL_MINTS.get(prin_mint) or meta.get(prin_mint) or {}).get("decimals")
                       or _solana_mint_decimals(prin_mint))
        col_dec  = int((KNOWN_SOL_MINTS.get(col_mint)  or meta.get(col_mint)  or {}).get("decimals")
                       or _solana_mint_decimals(col_mint))

        # Prices: Jupiter API → event fallback → USDC hardcode
        prin_price = float((price_data.get(prin_mint) or {}).get("price") or
                           event_prices.get(prin_mint) or
                           (1.0 if prin_mint == USDC_MINT else 0.0))
        col_price  = float((price_data.get(col_mint)  or {}).get("price") or
                           event_prices.get(col_mint) or 0.0)

        # Symbols: known-mints table → Jupiter meta → mint abbreviation
        prin_sym = ((KNOWN_SOL_MINTS.get(prin_mint) or meta.get(prin_mint) or {}).get("symbol") or "USDC").upper()
        col_sym  = ((KNOWN_SOL_MINTS.get(col_mint)  or meta.get(col_mint)  or {}).get("symbol") or col_mint[:6]).upper()

        # USD amounts
        prin_amount = prin_raw / (10 ** prin_dec)
        col_amount  = col_raw  / (10 ** col_dec)
        prin_usd    = prin_amount * prin_price
        col_usd     = col_amount  * col_price
        net_usd     = col_usd - prin_usd

        # Health: % distance to liquidation
        # lqtRatios scale = 1,000,000 (e.g. 900000 → 0.90 → 90% liquidation LTV)
        health_pct  = None
        health_rate = None
        lqt_ratios  = ledger.get("lqtRatios") or []
        if lqt_ratios and col_usd > 0 and prin_usd > 0:
            liq_raw = lqt_ratios[0]
            liq_ltv = liq_raw / 1_000_000 if liq_raw > 1 else float(liq_raw)
            cur_ltv = prin_usd / col_usd
            if 0 < cur_ltv < liq_ltv:
                health_pct  = round((liq_ltv / cur_ltv - 1) * 100, 1)
                health_rate = round(1.0 + health_pct / 100, 4)

        positions.append({
            "protocol":   "Loopscale",
            "chain":      "Solana",
            "chainId":    "sol",
            "type":       "Lending",
            "rabbyType":  "Lending",
            "healthRate": health_rate,
            "healthPct":  health_pct,
            "supplied":   [{"token": col_sym,  "amount": col_amount,  "usd": col_usd}],
            "borrowed":   [{"token": prin_sym, "amount": prin_amount, "usd": prin_usd}],
            "rewards":    [],
            "description": f"Loan {loan_id}",
            "_supSum":  col_usd,
            "_borSum":  prin_usd,
            "_rewSum":  0.0,
            "_net":     net_usd,
            "_poolId":  loan_id,
            "_source":  "loopscale-api",
        })
        total_net += net_usd
        print(f"[refresh]   Loan {loan_id}: "
              f"col={col_sym} ${col_usd:,.0f}  bor={prin_sym} ${prin_usd:,.0f}  "
              f"net=${net_usd:,.0f}  health={health_pct}%")

    return positions, total_net


def fetch_sol_defi_positions() -> tuple[list[dict], float]:
    """Primary entry-point for Solana DeFi positions.

    Tries REST APIs first (Loopscale). Falls back to headless Playwright scrape
    only if REST APIs return nothing.

    Returns (positions, total_sol_net_usd).
    """
    positions: list[dict] = []
    total_net = 0.0

    # Shared token metadata (avoids duplicate 20 MB download)
    try:
        meta = _jup_token_meta()
    except Exception:
        meta = {}

    # ── Loopscale REST API ──────────────────────────────────────────
    try:
        ls_pos, ls_net = loopscale_positions(meta=meta)
        positions.extend(ls_pos)
        total_net += ls_net
    except Exception as e:
        print(f"[refresh] Loopscale REST API failed: {e}", file=sys.stderr)

    if positions:
        print(f"[refresh] Solana DeFi via REST API: {len(positions)} positions, net=${total_net:,.0f}")
        return positions, total_net

    # ── Playwright fallback ─────────────────────────────────────────
    print("[refresh] REST API returned no positions — falling back to Playwright scrape …")
    return scrape_sol_defi_positions()


def scrape_sol_defi_positions() -> tuple[list[dict], float]:
    """Headless-browser scrape of Jupiter portfolio page → (positions, sol_net_worth)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("[refresh] playwright not installed — skipping Solana DeFi positions", file=sys.stderr)
        return [], 0.0

    url = f"https://jup.ag/portfolio/{SOL_WALLET}"
    print(f"[refresh] Opening {url} …")
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=(
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ))
        page = ctx.new_page()
        page.goto(url, wait_until="networkidle", timeout=60_000)
        # Wait for position data to appear
        try:
            page.wait_for_selector("text=Loopscale", timeout=30_000)
        except Exception:
            pass
        try:
            page.wait_for_selector("text=Kamino", timeout=10_000)
        except Exception:
            pass
        text = page.inner_text("body")
        browser.close()

    positions, sol_net = _parse_jupiter_text(text)
    print(f"[refresh] Jupiter: {len(positions)} Solana DeFi positions, sol_net=${sol_net:,.2f}")
    return positions, sol_net


# ────────────────────────── main ──────────────────────────

def main():
    ap = argparse.ArgumentParser()
    _root = Path(__file__).parent.parent  # repo root
    ap.add_argument("--xlsx", default=str(_root / "outputs/debank_positions_log.xlsx"))
    ap.add_argument("--snapshots-dir", default=str(_root / "outputs/snapshots"))
    ap.add_argument("--dashboard", default=str(_root / "docs/index.html"))
    ap.add_argument("--out", help="Optional: write the latest snapshot JSON to this path too")
    args = ap.parse_args()

    snapshot_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    try:
        reported_total = rabby_total_balance()
    except Exception as e:
        print(f"[refresh] total_balance fetch failed: {e}", file=sys.stderr)
        reported_total = None

    try:
        protocols = rabby_protocols()
    except Exception as e:
        print(f"[refresh] complex_protocol_list fetch failed: {e}", file=sys.stderr)
        return 1

    positions = normalize(protocols)
    print(f"[refresh] Rabby returned {len(positions)} raw positions across {len({(p['protocol'], p['chain']) for p in positions})} protocol/chain pairs")
    positions = [p for p in positions if keep(p)]
    print(f"[refresh] {len(positions)} positions after $50 filter")

    try:
        all_tokens = rabby_tokens()
    except Exception as e:
        print(f"[refresh] token_list fetch failed: {e}", file=sys.stderr)
        all_tokens = []
    pos_token_ids = collect_position_token_ids(protocols)
    wallet_toks = wallet_tokens(all_tokens, pos_token_ids)
    print(f"[refresh] {len(wallet_toks)} EVM wallet tokens over ${FILTER_USD:.0f}")

    try:
        sol_toks = solana_wallet_tokens()
        print(f"[refresh] {len(sol_toks)} Solana wallet tokens over ${FILTER_USD:.0f}")
        wallet_toks = wallet_toks + sol_toks
    except Exception as e:
        print(f"[refresh] Solana fetch failed: {e}", file=sys.stderr)

    sol_defi_positions, sol_net_worth = [], 0.0
    try:
        sol_defi_positions, sol_net_worth = fetch_sol_defi_positions()
        positions = positions + sol_defi_positions
    except Exception as e:
        print(f"[refresh] Jupiter scrape failed: {e}", file=sys.stderr)
        # Fall back to last known Solana positions from previous snapshot
        try:
            _snap_dir = Path(args.snapshots_dir)
            _prev_snaps = sorted(f for f in _snap_dir.glob("2*.json"))
            if _prev_snaps:
                _prev = json.loads(_prev_snaps[-1].read_text())
                _sol_pos = [p for p in _prev.get("positions", []) if p.get("chainId") == "sol"]
                if _sol_pos:
                    # Convert snapshot format back to internal format
                    for p in _sol_pos:
                        p["_supSum"] = p.get("supSum", 0)
                        p["_borSum"] = p.get("borSum", 0)
                        p["_rewSum"] = p.get("rewSum", 0)
                        p["_net"]    = p.get("net", 0)
                        p["_poolId"] = p.get("poolId", "")
                        p["_source"] = "jupiter-cached"
                    positions = positions + _sol_pos
                    sol_net_worth = sum(p["_net"] for p in _sol_pos)
                    print(f"[refresh] Using {len(_sol_pos)} cached Solana positions (sol_net=${sol_net_worth:,.0f})")
        except Exception as e2:
            print(f"[refresh] Solana cache fallback failed: {e2}", file=sys.stderr)

    sup = sum(p["_supSum"] for p in positions)
    bor = sum(p["_borSum"] for p in positions)
    hr_vals = [p["healthRate"] for p in positions if p.get("healthRate") is not None and p["healthRate"] > 0]
    totals = dict(
        sup=sup, bor=bor, net=sup - bor, count=len(positions),
        lowest_hr=min(hr_vals) if hr_vals else None,
    )

    source_label = "Rabby API + Jupiter"

    # Combine EVM reported total with Solana net worth for grand total
    if reported_total is not None and sol_net_worth > 0:
        combined_total = reported_total + sol_net_worth
    elif sol_net_worth > 0:
        combined_total = totals["net"] + sol_net_worth
    else:
        combined_total = reported_total

    xlsx_path = Path(args.xlsx)
    new_wb = append_to_xlsx(xlsx_path, snapshot_time, source_label, positions, totals, combined_total)
    print(f"[refresh] {'Created' if new_wb else 'Appended to'} {xlsx_path}")

    snap = snapshot_dict(positions, totals, combined_total, snapshot_time, source_label, wallet_toks=wallet_toks)

    snapshots_dir = Path(args.snapshots_dir)
    snap_path = save_timestamped_snapshot(snapshots_dir, snap)
    print(f"[refresh] Saved snapshot to {snap_path}")

    render_and_write_dashboard(snapshots_dir, Path(args.dashboard), snap)

    if args.out:
        Path(args.out).parent.mkdir(parents=True, exist_ok=True)
        with open(args.out, "w") as f:
            json.dump(snap, f, indent=2, default=str)
        print(f"[refresh] Also wrote latest snapshot JSON to {args.out}")

    print(f"[refresh] Net=${totals['net']:,.0f}  Supplied=${sup:,.0f}  Borrowed=${bor:,.0f}  Positions={totals['count']}  LowestHR={totals['lowest_hr']}  Reported=${reported_total:,.0f}" if reported_total else
          f"[refresh] Net=${totals['net']:,.0f}  Supplied=${sup:,.0f}  Borrowed=${bor:,.0f}  Positions={totals['count']}  LowestHR={totals['lowest_hr']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
