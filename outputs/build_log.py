"""Seed the DeBank positions log .xlsx with snapshot #1 (manual paste from debank.com)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

WALLET = "0xe16be042f9433779909a972669be3a2003956348"
SNAPSHOT_TIME = "2026-04-26 06:40:00"
SOURCE = "DeBank (manual snapshot)"
REPORTED_TOTAL = 619236.0

# Same data structure as the dashboard, parsed from the user's paste
positions = [
    # Morpho (Ethereum)
    dict(protocol="Morpho", chain="Ethereum", type="Lending", hr=2.73,
         supplied=[("mF-ONE", 148000.0, 159720.17)],
         borrowed=[("USDC", 49493.8163, 49483.92)]),
    dict(protocol="Morpho", chain="Ethereum", type="Lending", hr=1.30,
         supplied=[("PT-apyUSD-18JUN2026", 201006.0180, 196719.99)],
         borrowed=[("USDC", 130083.4389, 130057.43)]),
    dict(protocol="Morpho", chain="Ethereum", type="Lending", hr=1.11,
         supplied=[("PT-srNUSD-28MAY2026", 283979.2379, 281923.78)],
         borrowed=[("USDC", 233038.4910, 232991.89)]),
    # Fluid (Ethereum)
    dict(protocol="Fluid", chain="Ethereum", type="Lending", hr=1.02,
         supplied=[("USDT", 600005.6984, 600191.70), ("sUSDe", 339463.1684, 416836.52)],
         borrowed=[("USDT", 933537.3689, 933826.77)]),
    dict(protocol="Fluid", chain="Ethereum", type="Lending", hr=1.03,
         supplied=[("USDC", 462226.8746, 462134.45), ("syrupUSDC", 81271.6000, 94369.31)],
         borrowed=[("USDC", 497327.8050, 497228.36)]),
    dict(protocol="Fluid", chain="Ethereum", type="Lending", hr=1.11,
         supplied=[("reUSD", 131159.6529, 140790.57)],
         borrowed=[("GHO", 114065.1326, 113962.47)]),
    dict(protocol="Fluid", chain="Ethereum", type="Lending", hr=1.06,
         supplied=[("GHO", 104504.4351, 104410.38), ("USDC", 9556.1092, 9554.20)],
         borrowed=[("USDC", 93518.2572, 93499.56), ("GHO", 8559.2254, 8551.52)]),
    # Metronome (Ethereum)
    dict(protocol="Metronome", chain="Ethereum", type="Lending", hr=1.08,
         supplied=[("USDC", 364811.6013, 364738.65)],
         borrowed=[("msUSD", 287940.6864, 287100.12)]),
    # Curve LlamaLend (Ethereum)
    dict(protocol="Curve LlamaLend", chain="Ethereum", type="Lending", hr=1.04,
         supplied=[("sUSDe", 196098.2016, 240794.58)],
         borrowed=[("crvUSD", 217489.0200, 217453.18)]),
    dict(protocol="Curve LlamaLend", chain="Ethereum", type="Lending", hr=1.04,
         supplied=[("sfrxUSD", 275773.4355, 328537.81)],
         borrowed=[("crvUSD", 306325.9039, 306275.43)]),
    # Aave V3 (Ethereum)
    dict(protocol="Aave V3", chain="Ethereum", type="Lending", hr=1.05,
         supplied=[("USDT", 68010.5298, 68031.61)],
         borrowed=[("USDC", 50516.0026, 50505.90)]),
    # Silo (Ethereum)
    dict(protocol="Silo", chain="Ethereum", type="Lending", hr=1.26,
         supplied=[("PT-reUSDe-25JUN2026", 35130.2709, 34309.00)],
         borrowed=[("USDC", 25072.9422, 25067.93)]),
    # Fluid (Plasma)
    dict(protocol="Fluid", chain="Plasma", type="Lending", hr=1.04,
         supplied=[("sUSDai", 249570.0181, 269799.61)],
         borrowed=[("USDT0", 232920.8076, 232993.01)]),
    # Yuzu Money (Plasma)
    dict(protocol="Yuzu Money", chain="Plasma", type="Yield", hr=None,
         supplied=[("USDT0 (Protection Pool Cooldown #26)", 15479.3060, 15484.10)],
         borrowed=[]),
    # Aave V3 (Mantle)
    dict(protocol="Aave V3", chain="Mantle", type="Lending", hr=1.05,
         supplied=[("syrupUSDT", 69284.0832, 77777.00)],
         borrowed=[("USDT0", 68005.0641, 68026.15)]),
]


def position_id(p):
    sup = "+".join(sorted(t.split(" ")[0] for t, _, _ in p["supplied"])) or "none"
    bor = "+".join(sorted(t.split(" ")[0] for t, _, _ in p["borrowed"])) or "none"
    return f"{p['protocol'].replace(' ', '')}-{p['chain']}-{sup}~{bor}"


def position_name(p):
    sup = "+".join(t.split(" ")[0] for t, _, _ in p["supplied"]) or "—"
    bor = "+".join(t.split(" ")[0] for t, _, _ in p["borrowed"]) or "(no debt)"
    return f"{sup} vs {bor}"


def fmt_tokens(rows):
    return "; ".join(f"{t} {amt:,.4f}" for t, amt, _ in rows) if rows else ""


# Filter: keep positions where supplied OR net > $50
def keep(p):
    sup = sum(u for _, _, u in p["supplied"])
    bor = sum(u for _, _, u in p["borrowed"])
    net = sup - bor
    return abs(net) > 50 or sup > 50


positions_kept = [p for p in positions if keep(p)]

# Build workbook
wb = Workbook()

# ============== Sheet 1: Positions Log ==============
ws = wb.active
ws.title = "Positions Log"

headers = [
    "Snapshot (UTC)", "Wallet", "Chain", "Protocol", "Type",
    "Position ID", "Position", "Health Rate",
    "Supplied Tokens", "Supplied USD",
    "Borrowed Tokens", "Borrowed USD",
    "Net USD", "LTV", "Source",
]
ws.append(headers)

for p in positions_kept:
    sup_usd = sum(u for _, _, u in p["supplied"])
    bor_usd = sum(u for _, _, u in p["borrowed"])
    row_idx = ws.max_row + 1
    ws.append([
        SNAPSHOT_TIME,
        WALLET,
        p["chain"],
        p["protocol"],
        p["type"],
        position_id(p),
        position_name(p),
        p["hr"] if p["hr"] is not None else "",
        fmt_tokens(p["supplied"]),
        sup_usd,
        fmt_tokens(p["borrowed"]),
        bor_usd,
        f"=J{row_idx}-L{row_idx}",  # Net = Supplied - Borrowed
        f"=IF(J{row_idx}=0,\"\",L{row_idx}/J{row_idx})",  # LTV
        SOURCE,
    ])

# Style header
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F2937")
center = Alignment(horizontal="center", vertical="center")
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center

# Number formats
usd_fmt = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
ltv_fmt = '0.0%;[Red]-0.0%;"-"'
hr_fmt = '0.00'
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=8).number_format = hr_fmt          # Health Rate
    ws.cell(row=r, column=10).number_format = usd_fmt        # Supplied USD
    ws.cell(row=r, column=12).number_format = usd_fmt        # Borrowed USD
    ws.cell(row=r, column=13).number_format = usd_fmt        # Net USD
    ws.cell(row=r, column=14).number_format = ltv_fmt        # LTV

# Conditional format on Health Rate column (H)
red_fill = PatternFill("solid", start_color="FDECEA")
amber_fill = PatternFill("solid", start_color="FFF4E0")
green_fill = PatternFill("solid", start_color="E8F5EC")
hr_range = f"H2:H{ws.max_row}"
ws.conditional_formatting.add(hr_range,
    CellIsRule(operator="lessThan", formula=["1.05"], fill=red_fill))
ws.conditional_formatting.add(hr_range,
    CellIsRule(operator="between", formula=["1.05", "1.15"], fill=amber_fill))
ws.conditional_formatting.add(hr_range,
    CellIsRule(operator="greaterThanOrEqual", formula=["1.15"], fill=green_fill))

# Column widths
widths = {
    "A": 19, "B": 44, "C": 11, "D": 18, "E": 9,
    "F": 36, "G": 32, "H": 11,
    "I": 38, "J": 14,
    "K": 38, "L": 14,
    "M": 14, "N": 8, "O": 26,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"

# Apply font to all cells
arial = Font(name="Arial", size=10)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        if not c.font.bold:
            c.font = arial

# ============== Sheet 2: Snapshot Summary ==============
ws2 = wb.create_sheet("Snapshot Summary")
sum_headers = [
    "Snapshot (UTC)", "Wallet", "Reported Total USD",
    "Computed Net Equity USD", "Total Supplied USD", "Total Borrowed USD",
    "Position Count", "Lowest Health Rate", "Source",
]
ws2.append(sum_headers)

# Compute summary stats for this snapshot directly. The scheduled task will
# append a row of computed values for each new snapshot rather than relying
# on formulas — avoids cross-version formula compatibility issues.
total_sup = sum(sum(u for _, _, u in p["supplied"]) for p in positions_kept)
total_bor = sum(sum(u for _, _, u in p["borrowed"]) for p in positions_kept)
total_net = total_sup - total_bor
hr_values = [p["hr"] for p in positions_kept if p["hr"] is not None and p["hr"] > 0]
lowest_hr = min(hr_values) if hr_values else ""

ws2.append([
    SNAPSHOT_TIME,
    WALLET,
    REPORTED_TOTAL,
    total_net,
    total_sup,
    total_bor,
    len(positions_kept),
    lowest_hr,
    SOURCE,
])

for col_idx in range(1, len(sum_headers) + 1):
    c = ws2.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center

for col, num_fmt in [("C", usd_fmt), ("D", usd_fmt), ("E", usd_fmt), ("F", usd_fmt), ("H", hr_fmt)]:
    ws2[f"{col}2"].number_format = num_fmt

widths2 = {"A": 19, "B": 44, "C": 18, "D": 22, "E": 18, "F": 18, "G": 14, "H": 16, "I": 26}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for c in row:
        if not c.font.bold:
            c.font = arial

# ============== Sheet 3: README ==============
ws3 = wb.create_sheet("README")
readme = [
    ("DeBank Positions Log", True),
    ("", False),
    (f"Wallet: {WALLET}", False),
    ("Source page: https://debank.com/profile/" + WALLET, False),
    ("", False),
    ("Schema:", True),
    ("  Sheet 'Positions Log' — one row per (position, snapshot). Append a new row each time the scheduled task runs.", False),
    ("  Sheet 'Snapshot Summary' — one row per snapshot. Pulls totals from 'Positions Log' via SUMIFS keyed on Snapshot timestamp.", False),
    ("", False),
    ("Filtering rule:", True),
    ("  Positions are kept when Supplied USD > $50 OR |Net USD| > $50.", False),
    ("  Wallet token balances under $50 are excluded entirely.", False),
    ("", False),
    ("Position ID:", True),
    ("  Stable cross-snapshot key built from protocol + chain + sorted supplied tokens + sorted borrowed tokens.", False),
    ("  Use this column to track a position's evolution over time (filter / pivot by Position ID).", False),
    ("", False),
    ("Health Rate color coding:", True),
    ("  < 1.05  → red (high liquidation risk)", False),
    ("  1.05–1.15 → amber (watch closely)", False),
    ("  ≥ 1.15  → green (safer)", False),
    ("", False),
    ("Refresh:", True),
    ("  The scheduled task (every 12 hours) appends new rows for the current snapshot — it does not overwrite history.", False),
]
for i, (text, bold) in enumerate(readme, start=1):
    c = ws3.cell(row=i, column=1, value=text)
    c.font = Font(name="Arial", size=11, bold=bold)
ws3.column_dimensions["A"].width = 130

OUT = "/sessions/vibrant-friendly-johnson/mnt/outputs/debank_positions_log.xlsx"
wb.save(OUT)
print("Saved:", OUT)
print("Positions kept:", len(positions_kept), "of", len(positions))
print("Total Supplied:", sum(sum(u for _, _, u in p["supplied"]) for p in positions_kept))
print("Total Borrowed:", sum(sum(u for _, _, u in p["borrowed"]) for p in positions_kept))
